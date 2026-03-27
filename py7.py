import tkinter as tk
from tkinter import ttk, messagebox
import os
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk

EXCEL_FILE = "student_registration_data.xlsx"
BG_IMAGE_FILE = "BG_IMAGE_FILE.jpeg"

def save_data():
    """Validates inputs and saves data to the Excel file."""
    name = name_entry.get()
    roll = roll_entry.get()
    mobile = mobile_entry.get()
    email = email_entry.get()
    address = address_text.get("1.0", "end-1c") 
    stream = stream_combobox.get()
    
    if not all([name, roll, mobile, email, address, stream]):
        messagebox.showerror("Error", "Please fill out all fields.")
        return
    
    if not term_var.get():
        messagebox.showerror("Error", "You must agree to the Terms and Conditions.")
        return

    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"
            ws.append(["Name", "Roll Number", "Mobile", "Email", "Address", "Stream"])
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        ws.append([name, roll, mobile, email, address, stream])
        wb.save(EXCEL_FILE)
        
        messagebox.showinfo("Success", f"Registration successful!\nData saved to {EXCEL_FILE}")
        clear_form()
        
    except PermissionError:
        messagebox.showerror("File Error", "Cannot save. Please make sure the Excel file is closed.")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {e}")

def clear_form():
    """Clears all input fields."""
    name_entry.delete(0, tk.END)
    roll_entry.delete(0, tk.END)
    mobile_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)
    address_text.delete("1.0", tk.END)
    stream_combobox.set('')
    term_var.set(0)

root = tk.Tk()
root.title("Student Registration Form")
root.geometry("600x700")
root.resizable(False, False)

try:
    bg_image = Image.open(BG_IMAGE_FILE)
    bg_image = bg_image.resize((600, 700))
    bg_photo = ImageTk.PhotoImage(bg_image)
    
    bg_label = tk.Label(root, image=bg_photo)
    bg_label.place(x=0, y=0, relwidth=1, relheight=1)
except FileNotFoundError:
    print(f"Warning: '{BG_IMAGE_FILE}' not found. Loading form with standard background.")

frame = tk.Frame(root, bg="#f0f0f0", bd=5, relief="ridge")
frame.place(relx=0.5, rely=0.5, anchor="center", width=450, height=580)

title_label = tk.Label(frame, text="Student Registration", font=("Arial", 18, "bold"), bg="#f0f0f0")
title_label.pack(pady=(10, 20))

grid_frame = tk.Frame(frame, bg="#f0f0f0")
grid_frame.pack(fill="x", padx=20)

tk.Label(grid_frame, text="Full Name:", bg="#f0f0f0", font=("Arial", 10)).grid(row=0, column=0, sticky="w", pady=10)
name_entry = tk.Entry(grid_frame, width=30, font=("Arial", 10))
name_entry.grid(row=0, column=1, pady=10)

tk.Label(grid_frame, text="Roll Number:", bg="#f0f0f0", font=("Arial", 10)).grid(row=1, column=0, sticky="w", pady=10)
roll_entry = tk.Entry(grid_frame, width=30, font=("Arial", 10))
roll_entry.grid(row=1, column=1, pady=10)

tk.Label(grid_frame, text="Mobile Number:", bg="#f0f0f0", font=("Arial", 10)).grid(row=2, column=0, sticky="w", pady=10)
mobile_entry = tk.Entry(grid_frame, width=30, font=("Arial", 10))
mobile_entry.grid(row=2, column=1, pady=10)

tk.Label(grid_frame, text="Email Address:", bg="#f0f0f0", font=("Arial", 10)).grid(row=3, column=0, sticky="w", pady=10)
email_entry = tk.Entry(grid_frame, width=30, font=("Arial", 10))
email_entry.grid(row=3, column=1, pady=10)

tk.Label(grid_frame, text="Stream:", bg="#f0f0f0", font=("Arial", 10)).grid(row=4, column=0, sticky="w", pady=10)
streams = ["Science", "Commerce", "Arts", "Engineering", "Medical", "Other"]
stream_combobox = ttk.Combobox(grid_frame, values=streams, width=27, state="readonly", font=("Arial", 10))
stream_combobox.grid(row=4, column=1, pady=10)

tk.Label(grid_frame, text="Address:", bg="#f0f0f0", font=("Arial", 10)).grid(row=5, column=0, sticky="nw", pady=10)
address_text = tk.Text(grid_frame, width=30, height=4, font=("Arial", 10))
address_text.grid(row=5, column=1, pady=10)

term_var = tk.IntVar()
term_check = tk.Checkbutton(frame, text="I agree to the Terms and Conditions", variable=term_var, bg="#f0f0f0", font=("Arial", 9))
term_check.pack(pady=(15, 5))

submit_btn = tk.Button(frame, text="Register Student", font=("Arial", 12, "bold"), bg="#4CAF50", fg="white", width=20, command=save_data)
submit_btn.pack(pady=20)

root.mainloop()